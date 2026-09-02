# Le but de ce script est d'extraire les informations d'un document BCIAT
# Pour remplir la table Approvisionnement.

import sys
import os
import json
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
import csv

from provenance.reference_data import ReferenceData, load_reference_data, normalize
from provenance.transform_provenance_data import transform_provenance_data

### Configurations

## Workbook structure
SHEET_NAME = "Fournisseurs"
HEADER_COLUMN_FOURNISSEURS = "Fournisseur"

# How far down to look for the header row. It sits on row 11, 16 or 18
# depending on which year's template the plan was written on.
MAX_ROW_HEADER_SEARCH = 40

## Result format
NAME_COL_DOCUMENT = "Excel Ademe"

NAME_COL_ROW = "Ligne Excel"
NAME_COL_FOURNISSEUR = "Fournisseur"
NAME_COL_RESSOURCE = "Ressource"
NAME_COL_TONNAGE = "Tonnage"
NAME_COL_PROVENANCE = "Provenance"
NAME_COL_DATA_CONFIDENCE = "Niveau de confiance"
NAME_COL_RAW_PROVENANCE = 'Valeur brute de la colonne "Répartition approximative du combustible par département"'

# How each column we need announces itself, as a prefix of its normalized
# header. Columns are found by name rather than by index because the templates
# do not agree on either: the tonnage sits on column 3, 4 or 6, and the
# provenance on column 10, 14 or 21. The wording drifts too, hence a prefix —
# "Tonnage / an" in 2021, "Tonnage (t/an)" from 2023 on.
COLUMN_HEADER_PREFIXES = {
    NAME_COL_FOURNISSEUR: "fournisseur",
    NAME_COL_RESSOURCE: "sous categorie",
    NAME_COL_TONNAGE: "tonnage",
    NAME_COL_PROVENANCE: "repartition approximative",
}


class MyError(Exception):
    pass


def check_path_exists(file_path: str) -> None:
    if not os.path.exists(file_path):
        raise MyError(f"The file does not exist: {file_path}")


def check_file_is_xlsx(file_path: str) -> None:
    # splitext returns a tuple of root and extension
    file_extension = os.path.splitext(file_path)[1]

    if file_extension != ".xlsx":
        raise MyError(f"The file is not a xlsx file: {file_path}")


def find_index_row_headers(ws: Worksheet) -> int:
    """The row carrying "Fournisseur" in column A.

    Only an exact match counts: the sheet is titled "Fournisseurs" on row 1 and
    talks about a "Vérification tonnage Fournisseur" above the table, and
    neither of those is the header.
    """
    rows = ws.iter_rows(
        min_row=1, max_row=MAX_ROW_HEADER_SEARCH, min_col=1, max_col=1, values_only=True
    )

    for index, (value,) in enumerate(rows, start=1):
        if isinstance(value, str) and value.strip() == HEADER_COLUMN_FOURNISSEURS:
            return index

    raise MyError(
        f'No row in the first {MAX_ROW_HEADER_SEARCH} of the sheet "{SHEET_NAME}" '
        f'carries the header "{HEADER_COLUMN_FOURNISSEURS}" in column A.'
    )


def find_columns(ws: Worksheet, index_row_headers: int) -> dict[str, int]:
    """Where each column we need sits, read off the header row.

    The leftmost match wins, which is what tells the "Fournisseur" column from
    the "Fournisseur certifié PEFC/FSC" one that some templates add later on.
    """
    headers = next(
        ws.iter_rows(
            min_row=index_row_headers, max_row=index_row_headers, values_only=True
        )
    )
    normalized = [
        normalize(str(header)) if header is not None else "" for header in headers
    ]

    columns = {}
    for name, prefix in COLUMN_HEADER_PREFIXES.items():
        for index, header in enumerate(normalized):
            if header.startswith(prefix):
                columns[name] = index
                break

    missing = [name for name in COLUMN_HEADER_PREFIXES if name not in columns]
    if missing:
        raise MyError(
            f'The header row {index_row_headers} of the sheet "{SHEET_NAME}" has no '
            f"column for: {', '.join(missing)}."
        )

    return columns


def check_workbook_structure(wb: Workbook) -> None:
    if SHEET_NAME not in wb.sheetnames:
        raise MyError(f'The sheet "{SHEET_NAME}" does not exist in the file.')

    ws = wb[SHEET_NAME]
    find_columns(ws, find_index_row_headers(ws))


def find_index_last_row_data(ws: Worksheet, index_row_headers: int) -> int:
    # The last row of data is the row where the first Column ("Fournisseur") is None
    rows = ws.iter_rows(
        min_row=index_row_headers + 1, min_col=1, max_col=1, values_only=True
    )

    for offset, (value,) in enumerate(rows):
        if value is None:
            return index_row_headers + offset

    return ws.max_row


def extract_data_from_worksheet(
    wb_name: str, ws: Worksheet, reference_data: ReferenceData
) -> list[dict[str, str]]:
    extract_data = []

    index_row_headers = find_index_row_headers(ws)
    columns = find_columns(ws, index_row_headers)
    index_last_row = find_index_last_row_data(ws, index_row_headers)
    index_first_row = index_row_headers + 1
    rows = ws.iter_rows(
        min_row=index_first_row, max_row=index_last_row, values_only=True
    )
    for offset, row in enumerate(rows):
        raw_provenance = row[columns[NAME_COL_PROVENANCE]]
        provenance = transform_provenance_data(raw_provenance, reference_data)
        extract_data.append(
            {
                NAME_COL_DOCUMENT: wb_name,
                NAME_COL_ROW: index_first_row + offset,
                NAME_COL_FOURNISSEUR: row[columns[NAME_COL_FOURNISSEUR]],
                NAME_COL_RESSOURCE: row[columns[NAME_COL_RESSOURCE]],
                NAME_COL_TONNAGE: row[columns[NAME_COL_TONNAGE]],
                NAME_COL_PROVENANCE: {
                    "distribution": provenance["distribution"],
                    "unrecognized": provenance["unrecognized"],
                },
                NAME_COL_DATA_CONFIDENCE: provenance["status"],
                NAME_COL_RAW_PROVENANCE: raw_provenance,
            }
        )

    return extract_data


def main(argv: list[str]) -> int:
    if len(argv) != 1:
        print("Missing argument: you need to provide the path of a file.")
        print("Example: python3 bciat_bcib.py example.xlsx")
        return 2
    path = argv[0]

    try:
        print("Checking document...")
        check_path_exists(path)
        check_file_is_xlsx(path)
        wb = load_workbook(path, read_only=True, data_only=True)
        check_workbook_structure(wb)
        ws = wb[SHEET_NAME]
        print("✅ Document's structure is correct.")
        wb_name = os.path.basename(path)
        print("Extract and transform data...")
        data = extract_data_from_worksheet(wb_name, ws, load_reference_data())
        print("✅ Data extracted and transformed.")

        print("Writing results in csv file...")
        header = list(data[0].keys())
        rows = [list(row.values()) for row in data]
        header_and_rows = [header] + rows
        ouputFilePath = "import_" + wb_name + ".csv"
        with open(ouputFilePath, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(header_and_rows)
        print("✅ Results are written in some.csv.")

    except MyError as e:
        print(e, file=sys.stderr)
        return 2

    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
